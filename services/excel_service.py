import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter

class ExcelService:
    def __init__(self):
        # Professional Color Scheme (Matching the web UI)
        self.colors = {
            'header_bg': '1E40AF',      # Blue-800
            'header_text': 'FFFFFF',    # White
            'subheader_bg': '3B82F6',   # Blue-500
            'subheader_text': 'FFFFFF', # White
            'row_even': 'F8FAFC',       # Slate-50
            'row_odd': 'FFFFFF',        # White
            'input_bg': 'FEF3C7',       # Amber-100 (Editable)
            'input_border': 'F59E0B',   # Amber-500
            'total_bg': 'F0FDF4',       # Green-50
            'total_text': '064E3B'      # Green-900
        }
        
        # Styles
        self.border_thin = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        self.border_medium = Border(
            bottom=Side(style='medium', color='1E40AF')
        )
        
        self.font_header = Font(name='Calibri', size=12, bold=True, color=self.colors['header_text'])
        self.font_subheader = Font(name='Calibri', size=11, bold=True, color=self.colors['subheader_text'])
        self.font_normal = Font(name='Calibri', size=11, color='1E293B')
        self.font_bold = Font(name='Calibri', size=11, bold=True, color='1E293B')
        self.font_total = Font(name='Calibri', size=12, bold=True, color=self.colors['total_text'])

    def generate_project_excel(self, project_data):
        """Generate a complete Excel workbook for the project."""
        wb = openpyxl.Workbook()
        
        # 0. Pre-calculate dynamic columns (Accessory Names & Optional Items)
        all_accessory_names = set()
        all_optional_names = set()
        
        for fan in project_data.get('fans', []):
            weights = fan.get('weights', {})
            specs = fan.get('specifications', {})
            
            # Accessories
            if 'accessory_weight_details' in weights and weights['accessory_weight_details']:
                all_accessory_names.update(weights['accessory_weight_details'].keys())
            if 'custom_accessories' in specs and specs['custom_accessories']:
                all_accessory_names.update(specs['custom_accessories'].keys())
                
            # Optional Items
            if 'optional_items' in specs and specs['optional_items']:
                all_optional_names.update(specs['optional_items'].keys())
        
        sorted_acc_names = sorted(list(all_accessory_names))
        sorted_opt_names = sorted(list(all_optional_names))
        
        # Create sheets
        ws_quote = wb.active
        ws_quote.title = "Quotation"
        ws_specs = wb.create_sheet("Detailed Technical Specs")
        ws_internal = wb.create_sheet("Internal Costing")
        
        # Create Internal Costing Sheet first to get row mappings
        attr_row_map = self._create_costing_sheet(ws_internal, project_data, sorted_acc_names, sorted_opt_names)
        
        # 2. Detailed Technical Specs
        self._create_specs_sheet(ws_specs, project_data)
        
        # 1. Quotation Sheet (Client Facing)
        self._create_quotation_sheet(ws_quote, project_data, wb, attr_row_map) # Pass wb for named ranges, map for linking
        
        return wb

    def _create_quotation_sheet(self, ws, project, wb, attr_row_map):
        ws.title = "Quotation"
        ws.sheet_view.showGridLines = False
        ws.protection.sheet = False # Explicitly unlock
        
        # --- Header Section ---
        ws.merge_cells('B2:H2')
        cell = ws['B2']
        cell.value = "TCF Fan Pricing Tool v7.0 - Project Quotation"
        cell.font = Font(name='Calibri', size=16, bold=True, color=self.colors['header_bg'])
        cell.alignment = Alignment(horizontal='left')
        
        # Project Details Table
        details = [
            ("Enquiry Number:", project.get('enquiry_number')),
            ("Customer Name:", project.get('customer_name')),
            ("Sales Engineer:", project.get('sales_engineer')),
            ("Date:", project.get('created_at'))
        ]
        
        row = 4
        for label, value in details:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = self.font_bold
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = self.font_normal
            row += 1
            
            
        # --- Instructions ---
        ws['F4'] = "NOTE: To adjust prices, edit 'Fab Margin %' or 'BO Margin %' in the 'Internal Costing' sheet."
        ws['F4'].font = Font(color='DC2626', italic=True, size=10) # Red text
        ws.merge_cells('F4:K4')
        
        # --- Main Quotation Table ---
        row = 9
        headers = ["#", "Fan Model", "Tag", "Air Flow (CMH)", "Static Pressure (mmwc)", "Size", "Class", "Arrangement", "Qty", "Unit Price (₹)", "Total Price (₹)"]
        col_widths = [5, 20, 15, 15, 15, 10, 10, 15, 8, 20, 20]
        
        # Header Row
        for col, (header, width) in enumerate(zip(headers, col_widths), start=2):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.font_header
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
            ws.column_dimensions[get_column_letter(col)].width = width
            
        row += 1
        start_data_row = row
        
        # Get/Verify Row Indices from Map
        # We need: Fab Cost, Total BO, Fab Margin %, BO Margin %
        # Default fallback if not found (should not happen if Costing is created first)
        row_fab_cost = attr_row_map.get("Fabrication Cost (₹)", 30)
        row_bo_cost = attr_row_map.get("Total Bought Out (₹)", 31)
        row_fab_margin = attr_row_map.get("Fab Margin %", 35)
        row_bo_margin = attr_row_map.get("BO Margin %", 36)
        
        for idx, fan in enumerate(project.get('fans', []), 1):
            specs = fan.get('specifications', {})
            
            # Base data
            ws.cell(row=row, column=2, value=idx).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=specs.get('Fan Model'))
            ws.cell(row=row, column=4, value=specs.get('fan_tag', '-'))
            ws.cell(row=row, column=5, value=specs.get('air_flow', '-'))
            ws.cell(row=row, column=6, value=specs.get('static_pressure', '-'))
            ws.cell(row=row, column=7, value=specs.get('Fan Size'))
            ws.cell(row=row, column=8, value=specs.get('Class'))
            ws.cell(row=row, column=9, value=specs.get('Arrangement'))
            ws.cell(row=row, column=10, value=1).alignment = Alignment(horizontal='center')
            
            cost_sheet = "Internal Costing"
            
            # Fan Data Column: Fan 1 -> Col B (2), Fan 2 -> Col C (3)
            fan_col_letter = get_column_letter(1 + idx)
            
            # Linking Formula: Use Cost References AND Margin References from Internal Costing
            ref_fab_cost = f"'{cost_sheet}'!{fan_col_letter}{row_fab_cost}"
            ref_bo_cost = f"'{cost_sheet}'!{fan_col_letter}{row_bo_cost}"
            ref_fab_margin = f"'{cost_sheet}'!{fan_col_letter}{row_fab_margin}"
            ref_bo_margin = f"'{cost_sheet}'!{fan_col_letter}{row_bo_margin}"
            
            # Margin Logic: Handle 25 (integer) vs 0.25 (decimal)
            # Formula: IF(Margin>1, Margin/100, Margin)
            margin_fab_calc = f"IF({ref_fab_margin}>1, {ref_fab_margin}/100, {ref_fab_margin})"
            margin_bo_calc = f"IF({ref_bo_margin}>1, {ref_bo_margin}/100, {ref_bo_margin})"
            
            # Price = (FabCost / (1 - FabMargin)) + (BOCost / (1 - BOMargin))
            formula = f"=({ref_fab_cost}/(1-{margin_fab_calc})) + ({ref_bo_cost}/(1-{margin_bo_calc}))"
            
            price_cell = ws.cell(row=row, column=11)
            price_cell.value = formula
            price_cell.number_format = '₹ #,##0.00'
            
            # Total Formula (= Unit Price * Qty)
            total_cell = ws.cell(row=row, column=12)
            total_cell.value = f"=K{row}*J{row}"
            total_cell.number_format = '₹ #,##0.00'
            
            # Apply styling
            for col in range(2, 13):
                c = ws.cell(row=row, column=col)
                c.border = self.border_thin
                c.font = self.font_normal
                
            row += 1
            
        # --- Totals Row ---
        last_data_row = row - 1
        ws.cell(row=row, column=2, value="TOTAL Project Value").font = self.font_total
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        
        grand_total_cell = ws.cell(row=row, column=12)
        grand_total_cell.value = f"=SUM(L{start_data_row}:L{last_data_row})"
        grand_total_cell.font = self.font_total
        grand_total_cell.number_format = '₹ #,##0.00'
        grand_total_cell.fill = PatternFill(start_color=self.colors['total_bg'], end_color=self.colors['total_bg'], fill_type='solid')

    def _create_specs_sheet(self, ws, project):
        # ws = wb.create_sheet("Detailed Technical Specs") # Removed, passed in
        ws.sheet_view.showGridLines = False
        
        headers = ["Fan #", "Model", "Tag", "Air Flow (CMH)", "Static Pressure (mmwc)", "Size", "Class", "Type", "Material", "Motor kW", "Motor Brand", "Drive Type", "Isolators", "Accessories"]
        col_widths = [8, 15, 15, 15, 15, 10, 10, 15, 15, 12, 12, 15, 15, 40]
        
        # Header Row
        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.font_subheader
            cell.fill = PatternFill(start_color=self.colors['subheader_bg'], end_color=self.colors['subheader_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = width
            
        row = 2
        for idx, fan in enumerate(project.get('fans', []), 1):
            specs = fan.get('specifications', {})
            motor = fan.get('motor', {})
            weights = fan.get('weights', {})
            
            # Accessories list
            acc_list = []
            # Standard
            if weights.get('accessory_weight_details'):
                 acc_list.extend(weights['accessory_weight_details'].keys())
            # Custom
            if specs.get('custom_accessories'):
                 acc_list.extend(specs['custom_accessories'].keys())
            # Optional
            if specs.get('optional_items'):
                 acc_list.extend(specs['optional_items'].keys())     
                 
            ws.cell(row=row, column=1, value=f"Fan {idx}")
            ws.cell(row=row, column=2, value=specs.get('Fan Model'))
            ws.cell(row=row, column=3, value=specs.get('fan_tag', '-'))
            ws.cell(row=row, column=4, value=specs.get('air_flow', '-'))
            ws.cell(row=row, column=5, value=specs.get('static_pressure', '-'))
            ws.cell(row=row, column=6, value=specs.get('Fan Size'))
            ws.cell(row=row, column=7, value=specs.get('Class'))
            ws.cell(row=row, column=8, value=specs.get('Arrangement'))
            ws.cell(row=row, column=9, value=specs.get('material'))
            ws.cell(row=row, column=10, value=motor.get('kw'))
            ws.cell(row=row, column=11, value=motor.get('brand'))
            ws.cell(row=row, column=12, value=specs.get('drive_pack'))
            ws.cell(row=row, column=13, value=specs.get('vibration_isolators'))
            ws.cell(row=row, column=14, value=", ".join(acc_list))
            
            row += 1

    def _create_costing_sheet(self, ws, project, sorted_acc_names, sorted_opt_names):
        # ws = wb.create_sheet("Internal Costing") # Removed, passed in
        ws.protection.sheet = False # Explicitly unlock
        
        # Instructions
        ws['A1'] = "INTERNAL USE ONLY - DETAILED OPERATIONAL DATA"
        ws['A1'].font = Font(color='DC2626', bold=True, size=14)
        
        # --- Define Sections ---
        # Format: (Section Title, [List of Attribute Names])
        
        # 1. Identity
        sec_identity = (None, ["Fan #", "Model", "Size", "Air Flow (CMH)", "Static Pressure (mmwc)"])
        
        # 2. Weights
        sec_weights_attrs = ["Bare Weight (kg)", "Standard Acc Wt (kg)", "Custom Acc Wt (kg)", "Total Weight (kg)", "MS Weight (kg)", "SS Weight (kg)"]
        
        sec_weights = ("--- WEIGHTS & ACCESSORIES ---", sec_weights_attrs + [f"{name} (kg)" for name in sorted_acc_names])
        
        # 3. Specs
        sec_specs = ("--- COMPONENT SPECIFICATIONS ---", [
            "Motor kW", "Motor Brand", "Motor Efficiency", "Motor Pole",
            "Bearing Brand", "Isolators Brand", # Requested but data might be missing
            "Isolators Qty", "Shaft Dia (mm)", "Drive Pack", "Material"
        ])
        
        # 4. Costs
        sec_costs = ("--- COMPONENT COSTS ---", [
            "Motor Price (₹)", "Bearing Price (₹)", "Drive Pack Price (₹)", "Isolator Price (₹)", 
            "Fabrication Cost (₹)", "Total Bought Out (₹)" 
            # Note: Optional Items Cost is separate, let's break it down
        ])
        
        # 5. Optional Items
        # Dynamic rows for optional items (Price/Details)
        # We will list them as "Opt: Item Name (₹)"
        sec_optional = ("--- OPTIONAL ITEMS ---", [f"{name} (₹)" for name in sorted_opt_names])
        
        # 6. Margins & Totals
        sec_totals = ("--- MARGINS & TOTALS ---", [
            "Fab Margin %", "BO Margin %",
            "Fab Selling Price (₹)", "BO Selling Price (₹)", "Total Selling Price (₹)"
        ])
        
        sections = [sec_identity, sec_weights, sec_specs, sec_costs, sec_optional, sec_totals]
        
        # Flatten for writing
        # Map: Attribute Name -> Row Index (for populating data)
        attr_row_map = {} 
        current_row = 3
        
        ws.column_dimensions['A'].width = 35
        
        for title, attrs in sections:
            if title:
                cell = ws.cell(row=current_row, column=1, value=title)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='64748B', end_color='64748B', fill_type='solid') # Slate-500
                cell.alignment = Alignment(horizontal='center')
                # Merge across some columns? dynamic length... lets just fill A
                current_row += 1
            
            for attr in attrs:
                cell = ws.cell(row=current_row, column=1, value=attr)
                cell.font = self.font_bold
                cell.border = self.border_thin
                cell.fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
                attr_row_map[attr] = current_row
                current_row += 1
                
        # Write Fan Data (Columns B, C, D...)
        for fan_idx, fan in enumerate(project.get('fans', []), 1):
            col = fan_idx + 1 # Start at Col B
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
            
            costs = fan.get('costs', {})
            weights = fan.get('weights', {})
            motor = fan.get('motor', {})
            specs = fan.get('specifications', {})
            
            def get_val(d, k): return d.get(k, 0) or 0
            
            # BRAND LOGIC
            # Use stored values from Fan Calculator
            bearing_brand = specs.get('bearing_brand', '')
            isolator_brand = specs.get('vibration_isolators', '')
            arrangement = str(specs.get('Arrangement', ''))
            
            # If Arrangement 4 (Direct), Bearing Brand is N/A
            if arrangement == '4':
                bearing_brand = "N/A"
            
            # Format Isolator Brand (e.g. 'polybond' -> 'Polybond')
            if isolator_brand == 'not_required':
                isolator_brand = "Not Required"
            elif isolator_brand:
                isolator_brand = isolator_brand.title()
                
            # Double check for empty/missing
            if not bearing_brand: bearing_brand = ""
            if not isolator_brand: isolator_brand = ""
            
            # Prepare Data Dict: Attribute Name -> Value
            data_map = {}
            
            # Identity
            data_map["Fan #"] = fan_idx
            data_map["Model"] = specs.get('Fan Model')
            data_map["Size"] = specs.get('Fan Size')
            data_map["Air Flow (CMH)"] = specs.get('air_flow', '-')
            data_map["Static Pressure (mmwc)"] = specs.get('static_pressure', '-')
            
            # Weights
            total_wt = get_val(weights, 'total_weight')
            data_map["Bare Weight (kg)"] = get_val(weights, 'bare_fan_weight')
            data_map["Standard Acc Wt (kg)"] = get_val(weights, 'accessory_weight')
            
            # Custom Acc Wt logic
            custom_acc_wt = total_wt - get_val(weights, 'bare_fan_weight') - get_val(weights, 'accessory_weight')
            if custom_acc_wt < 0: custom_acc_wt = 0
            data_map["Custom Acc Wt (kg)"] = custom_acc_wt
            data_map["Total Weight (kg)"] = total_wt
            
            # Mixed Material Breakdown
            material = specs.get('material', 'ms')
            if str(material).lower() == 'mixed':
                ms_percent = float(specs.get('ms_percentage', 0) or 0)
                ms_weight = total_wt * (ms_percent / 100.0)
                ss_weight = total_wt - ms_weight
                data_map["MS Weight (kg)"] = ms_weight
                data_map["SS Weight (kg)"] = ss_weight
            
            # Dynamic Accessories
            fan_acc_details = weights.get('accessory_weight_details', {}) or {}
            for name in sorted_acc_names:
                data_map[f"{name} (kg)"] = fan_acc_details.get(name, 0)
                
            # Specs
            data_map["Motor kW"] = motor.get('kw', '')
            data_map["Motor Brand"] = motor.get('brand', '')
            data_map["Motor Efficiency"] = motor.get('efficiency', '')
            data_map["Motor Pole"] = motor.get('pole', '')
            data_map["Bearing Brand"] = bearing_brand
            data_map["Isolators Brand"] = isolator_brand
            data_map["Isolators Qty"] = get_val(weights, 'no_of_isolators')
            data_map["Shaft Dia (mm)"] = get_val(weights, 'shaft_diameter')
            data_map["Drive Pack"] = specs.get('drive_pack')
            data_map["Material"] = material
            
            # Costs
            data_map["Motor Price (₹)"] = get_val(costs, 'discounted_motor_price')
            data_map["Bearing Price (₹)"] = get_val(costs, 'bearing_price')
            data_map["Drive Pack Price (₹)"] = get_val(costs, 'drive_pack_price')
            data_map["Isolator Price (₹)"] = get_val(costs, 'vibration_isolators_price')
            data_map["Fabrication Cost (₹)"] = get_val(costs, 'fabrication_cost')
            data_map["Total Bought Out (₹)"] = get_val(costs, 'bought_out_cost')
            
            # Optional Items
            # Need to get details. In routes.py we construct 'specifications' -> 'optional_items' (Dict Name->Price)
            opt_items = specs.get('optional_items', {}) or {}
            for name in sorted_opt_names:
                # Value matches name?
                # The stored format in specs['optional_items'] is {Name: Price}
                val = opt_items.get(name, 0)
                try:
                    val = float(val)
                except:
                    val = 0
                data_map[f"{name} (₹)"] = val
            
            # Margins
            data_map["Fab Margin %"] = specs.get('fabrication_margin', 25)
            data_map["BO Margin %"] = specs.get('bought_out_margin', 25)
            data_map["Fab Selling Price (₹)"] = get_val(costs, 'fabrication_selling_price')
            data_map["BO Selling Price (₹)"] = get_val(costs, 'bought_out_selling_price')
            data_map["Total Selling Price (₹)"] = get_val(costs, 'total_selling_price')
            
            # Populate Cells
            for attr, row_idx in attr_row_map.items():
                val = data_map.get(attr, "")
                cell = ws.cell(row=row_idx, column=col)
                
                # Dynamic Formulas for Selling Prices
                # We need column letter for current fan (e.g. B, C)
                fan_col_letter = get_column_letter(col)
                
                if attr == "Fab Selling Price (₹)":
                    # Formula: FabCost / (1 - FabMargin)
                    # Handle both Integer 25 and Decimal 0.25
                    row_cost = attr_row_map.get("Fabrication Cost (₹)")
                    row_margin = attr_row_map.get("Fab Margin %")
                    if row_cost and row_margin:
                        ref_cost = f"{fan_col_letter}{row_cost}"
                        ref_margin = f"{fan_col_letter}{row_margin}"
                        # IF(Margin>1, Margin/100, Margin) covers both cases.
                        # IF(Margin=1, 0) covers 100% margin edge case (DIV/0) -> but Excel handles error.
                        # Using IFERROR just in case
                        val = f"=IFERROR({ref_cost} / (1 - IF({ref_margin}>1, {ref_margin}/100, {ref_margin})), {ref_cost})"
                    
                elif attr == "BO Selling Price (₹)":
                    row_cost = attr_row_map.get("Total Bought Out (₹)")
                    row_margin = attr_row_map.get("BO Margin %")
                    if row_cost and row_margin:
                        ref_cost = f"{fan_col_letter}{row_cost}"
                        ref_margin = f"{fan_col_letter}{row_margin}"
                        val = f"=IFERROR({ref_cost} / (1 - IF({ref_margin}>1, {ref_margin}/100, {ref_margin})), {ref_cost})"

                elif attr == "Total Selling Price (₹)":
                    row_fab_price = attr_row_map.get("Fab Selling Price (₹)")
                    row_bo_price = attr_row_map.get("BO Selling Price (₹)")
                    if row_fab_price and row_bo_price:
                        ref_fab = f"{fan_col_letter}{row_fab_price}"
                        ref_bo = f"{fan_col_letter}{row_bo_price}"
                        val = f"={ref_fab} + {ref_bo}"

                cell.value = val
                cell.border = self.border_thin
                
                # Format
                if isinstance(val, (int, float)) or (isinstance(val, str) and val.startswith("=")):
                    if "Price" in attr or "Cost" in attr or "(₹)" in attr:
                        cell.number_format = '#,##0.00'
                    elif "Weight" in attr or "(kg)" in attr or "%" in attr or "Qty" in attr:
                         if "%" in attr:
                             if isinstance(val, (int, float)) and val > 1: 
                                 cell.value = val / 100.0
                             cell.number_format = '0.00%'
                         else:
                             cell.number_format = '#,##0'
        # UNLOCKED - Removed protection block
        # ws.protection.sheet = True ... removed
        
        return attr_row_map
 

    # UPDATE QUOTATION FORMULAS
    # We added dynamic columns, so Q and R are shifted!
    # Original Headers Count (Hardcoded): 23
    # Dynamic Headers Count: len(sorted_acc_names)
    # New "Total Bought Out" Index: 
    #   Prefix (3) + Dynamic (N) + Suffix_Index_of_BO
    #   Prefix: Fan#, Model, BareWt (3 columns)
    #   Dynamic: N
    #   Suffix: TotalAcc, TotalWt, MotorKW, Brand, IsoQty, Shaft, Drive, Mat, MotPrice, BrngPrice, DrvPrice, IsoPrice, OptPrice, TotalBO...
    #   TotalBO is at Suffix Index 13 (0-based) -> 14th item in Suffix
    #   FabCost is at Suffix Index 14 -> 15th item in Suffix
    
    # So Total BO Column Index = 2 (Start) + 3 (Prefix) + N + 13
    # = 18 + N
    # Fab Cost Column Index = 18 + N + 1
    # = 19 + N
    
    # We need to calculate this per project? 
    # Wait, the Quotation sheet is generated BEFORE or AFTER?
    # It's generated BEFORE in the code (`_create_quotation_sheet` called before `_create_costing_sheet`).
    # BUT `_create_quotation_sheet` needs to know the columns.
    # Refactor: Move dynamic column logic to `generate_project_excel` or make `_create_quotation_sheet` aware of keys.
    # Actually, iterate keys in `_create_quotation_sheet` too to find N.


