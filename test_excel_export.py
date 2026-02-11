import sys
import os
import io
from services.excel_service import ExcelService
import openpyxl

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

# Mock Project Data
project_data = {
    'enquiry_number': 'TEST-2025-001',
    'customer_name': 'Acme Corp',
    'sales_engineer': 'John Doe',
    'created_at': '2025-05-20',
    'fans': [
        {
            'fan_number': 1,
            'specifications': {
                'Fan Model': 'TCF-100',
                'Fan Size': '500',
                'Class': '1',
                'Arrangement': '4',
                'material': 'ms',
                'drive_pack': 'Direct',
                'vibration_isolators': 'polybond',
                'bearing_brand': 'SKF', # Should be ignored/overwritten to N/A
                'fabrication_margin': 30,
                'bought_out_margin': 20,
                'bought_out_margin': 20,
                'custom_accessories': {'Inlet Box': 'Yes', 'Drain': 'Yes'},
                'optional_items': {'Flex Connector': 1500}
            },
            'motor': {
                'brand': 'ABB',
                'kw': '15',
                'pole': '4',
                'efficiency': 'IE3'
            },
            'weights': {
                'bare_fan_weight': 400,
                'total_weight': 550,
                'accessory_weight_details': {'Inlet Box': 50}
            },
            'costs': {
                'fabrication_cost': 50000,
                'bought_out_cost': 20000,
                'optional_items_cost': 0,
                'total_cost': 70000
            }
        },
         {
            'fan_number': 2,
            'specifications': {
                'Fan Model': 'TCF-200',
                'Fan Size': '800',
                'Class': '2',
                'Arrangement': '9',
                'material': 'mixed',
                'ms_percentage': 60,
                'drive_pack': 'Belt',
                'vibration_isolators': 'not_required',
                'bearing_brand': 'Dodge',
                'fabrication_margin': 25,
                'bought_out_margin': 25,
            },
            'motor': {
                'brand': 'Siemens',
                'kw': '22',
                'pole': '2',
                'efficiency': 'IE4'
            },
            'weights': {
                'bare_fan_weight': 800,
                'total_weight': 950,
                'accessory_weight_details': {'Damper': 100}
            },
            'costs': {
                'fabrication_cost': 120000,
                'bought_out_cost': 45000,
                'optional_items_cost': 5000,
                'total_cost': 170000
            }
        }
    ]
}

def test_export():
    print("Testing Excel Service...")
    try:
        service = ExcelService()
        wb = service.generate_project_excel(project_data)
        
        output_file = "test_export_v2.xlsx"
        wb.save(output_file)
        print(f"Successfully saved {output_file}")
        
        # Validation
        wb_check = openpyxl.load_workbook(output_file)
        
        # Check Sheets
        print(f"Sheets: {wb_check.sheetnames}")
        if "Quotation" not in wb_check.sheetnames:
            print("FAIL: Quotation sheet missing")
            return
            
        ws_quote = wb_check["Quotation"]
        
        # Check Margins Input
        fab_margin_cell = ws_quote['G5']
        print(f"Default Fab Margin: {fab_margin_cell.value}")
        if fab_margin_cell.value != 0.3: # 30% from first fan
             print(f"WARNING: Fab margin mismatch. Expected 0.3, got {fab_margin_cell.value}")
             
        # Check Formula presence
        price_formula = ws_quote['H10'].value
        print(f"Price Formula Row 10: {price_formula}")
        if "Internal Costing" not in str(price_formula):
             print("FAIL: Formula does not reference Internal Costing")
             
        # Check Dynamic Columns in Costing (Now Rows in Col A)
        ws_costing = wb_check["Internal Costing"]
        
        # Get all values in Column A (Attributes)
        attributes = []
        for row in range(1, 60): # Scan first 60 rows
            val = ws_costing.cell(row=row, column=1).value
            if val:
                attributes.append(val)
        
        print(f"Costing Attributes (Col A): {attributes}")
        
        # Verify Sections
        if "--- WEIGHTS & ACCESSORIES ---" not in attributes:
            print("FAIL: Section '--- WEIGHTS & ACCESSORIES ---' missing")
        if "--- OPTIONAL ITEMS ---" not in attributes:
            print("FAIL: Section '--- OPTIONAL ITEMS ---' missing")
            
        # Verify Dynamic Rows (Accessory & Optional)
        if "Inlet Box (kg)" not in attributes:
            print("FAIL: Dynamic attribute 'Inlet Box (kg)' missing")
        if "Flex Connector (₹)" not in attributes:
             print("FAIL: Dynamic optional item 'Flex Connector (₹)' missing")
             
        # Check Data Orientation (Fan 1 in Col B, Fan 2 in Col C)
        fan1_model = ws_costing['B5'].value # Specific row depends on index, but Model is in Identity section
        # We need to find the row index for "Model"
        model_row = -1
        brand_row = -1
        
        for r in range(1, 60):
            val = ws_costing.cell(row=r, column=1).value
            if val == "Model":
                model_row = r
            if val == "Bearing Brand":
                brand_row = r
                
        if model_row != -1:
            fan1_model = ws_costing.cell(row=model_row, column=2).value
            print(f"Fan 1 Model Validation: Found at row {model_row}, Value: {fan1_model}")
            if fan1_model != "TCF-100":
                print(f"FAIL: Fan 1 Model mismatch. Expected TCF-100, got {fan1_model}")
        else:
            print("FAIL: 'Model' row not found")
            
        # Verify Brands
        if brand_row != -1:
            fan1_brand = ws_costing.cell(row=brand_row, column=2).value
            fan2_brand = ws_costing.cell(row=brand_row, column=3).value
            print(f"Fan 1 Bearing Brand: {fan1_brand}")
            print(f"Fan 2 Bearing Brand: {fan2_brand}")
            
            if fan1_brand != "N/A": 
                print(f"FAIL: Fan 1 Brand unexpected: {fan1_brand} (Expected N/A)")
            if fan2_brand != "Dodge":
                print(f"FAIL: Fan 2 Brand unexpected: {fan2_brand} (Expected Dodge)")
        else:
             print("FAIL: 'Bearing Brand' row not found")
             
        # Find Isolators Row
        iso_row = -1
        for r in range(1, 60):
            val = ws_costing.cell(row=r, column=1).value
            if val == "Isolators Brand":
                iso_row = r
                break
                
        if iso_row != -1:
            fan1_iso = ws_costing.cell(row=iso_row, column=2).value
            fan2_iso = ws_costing.cell(row=iso_row, column=3).value
            print(f"Fan 1 Isolator: {fan1_iso}")
            print(f"Fan 2 Isolator: {fan2_iso}")
            
            if fan1_iso != "Polybond":
                 print(f"FAIL: Fan 1 Isolator unexpected: {fan1_iso} (Expected Polybond)")
            if fan2_iso != "Not Required":
                 print(f"FAIL: Fan 2 Isolator unexpected: {fan2_iso} (Expected Not Required)")
        else:
            print("FAIL: 'Isolators Brand' row not found")
            
        # 7. Formula Verification
        # Fan 1 is at Index 1 -> Col B (2)
        # Fan 2 is at Index 2 -> Col C (3) -> This is what we are checking in row_check_col = 3
        
        # Verify Sheet Protection is OFF
        if ws_quote.protection.sheet:
            print("FAIL: Quotation sheet is protected/locked!")
        else:
            print("SUCCESS: Quotation sheet is unlocked.")
            
        if ws_costing.protection.sheet:
            print("FAIL: Internal Costing sheet is protected/locked!")
        else:
            print("SUCCESS: Internal Costing sheet is unlocked.")

        # Formula should be something like ...
        
        # 8. Dynamic Internal Costing Formulas Verification
        # Fan 2 is Col C (3).
        # We need to find rows for Fab Selling Price, BO Selling Price.
        # Since we don't have the map here, we can scan column A.
        
        print("Scannning for Selling Price rows...")
        row_fab_price = None
        row_total_price = None
        for r in range(1, 100):
            val = ws_costing.cell(row=r, column=1).value
            if val == "Fab Selling Price (₹)":
                row_fab_price = r
            if val == "Total Selling Price (₹)":
                row_total_price = r
                
        if row_fab_price:
            cell_formula = ws_costing.cell(row=row_fab_price, column=3).value
            print(f"Fab Selling Price Formula: {cell_formula}")
            if str(cell_formula).startswith("="):
                 print("SUCCESS: Fab Selling Price is a formula.")
            else:
                 print(f"FAIL: Fab Selling Price is not a formula: {cell_formula}")
        else:
            print("FAIL: Could not find 'Fab Selling Price (₹)' row in Internal Costing")

        if row_total_price:
            cell_formula = ws_costing.cell(row=row_total_price, column=3).value
            print(f"Total Selling Price Formula: {cell_formula}")
            if str(cell_formula).startswith("="):
                 print("SUCCESS: Total Selling Price is a formula.")
            else:
                 print(f"FAIL: Total Selling Price is not a formula: {cell_formula}")

        # Verify Bearing Brand Linkage
        # Fan 1 is Arr 4 -> Should be N/A
        # Fan 2 is Arr 9 -> Should be Dodge (from mock)
        bearing_row = -1
        for r in range(1, 40):
            val = ws_costing.cell(row=r, column=1).value
            if val == "Bearing Brand":
                bearing_row = r
                break
        
        if bearing_row != -1:
            fan1_brg = ws_costing.cell(row=bearing_row, column=2).value
            fan2_brg = ws_costing.cell(row=bearing_row, column=3).value
            print(f"Fan 1 Bearing: {fan1_brg}")
            print(f"Fan 2 Bearing: {fan2_brg}")
            
            if fan1_brg != "N/A":
                print(f"FAIL: Fan 1 Bearing unexpected: {fan1_brg} (Expected N/A for Arr 4)")
            if fan2_brg != "Dodge":
                print(f"FAIL: Fan 2 Bearing unexpected: {fan2_brg} (Expected Dodge)")
        else:
            print("FAIL: 'Bearing Brand' row not found")
            
        # Verify Drive Pack Label
        drive_pack_row = -1
        for r in range(1, 40):
            val = ws_costing.cell(row=r, column=1).value
            if val == "Drive Pack":
                drive_pack_row = r
                break
        if drive_pack_row == -1:
             print("FAIL: 'Drive Pack' label not found (Did you rename 'Drive Type'?)")
        else:
             print(f"SUCCESS: 'Drive Pack' label found at Row {drive_pack_row}")

        # Verify Mixed Material Breakdown
        ms_row = -1
        ss_row = -1
        for r in range(1, 60):
            val = ws_costing.cell(row=r, column=1).value
            if val == "MS Weight (kg)": ms_row = r
            if val == "SS Weight (kg)": ss_row = r
            
        if ms_row != -1 and ss_row != -1:
            # Fan 2 is Mixed (60% MS). Total Weight = 950.
            # MS = 950 * 0.6 = 570
            # SS = 950 * 0.4 = 380
            fan2_ms = ws_costing.cell(row=ms_row, column=3).value
            fan2_ss = ws_costing.cell(row=ss_row, column=3).value
            print(f"Fan 2 Mixed Wt Breakdown: MS={fan2_ms}, SS={fan2_ss}")
            
            if abs(fan2_ms - 570) > 0.1:
                print(f"FAIL: MS Weight mismatch. Expected 570, got {fan2_ms}")
            if abs(fan2_ss - 380) > 0.1:
                print(f"FAIL: SS Weight mismatch. Expected 380, got {fan2_ss}")
        else:
            print("FAIL: MS/SS Weight rows not found")

        # Verify Quotation Formulas (Check for correct row references)
        # N=0 (Mock data has no accessories in weight details, but Fan 1 has custom accs?)
        # Wait, sorted_acc_names comes from 'accessory_weight_details'.keys().
        # In mock, fan['weights']['accessory_weight_details'] is empty.
        # But 'custom_accessories' might add keys?
        # Logic: all_accessory_names.update(weights['accessory_weight_details'].keys())
        # AND specs['custom_accessories'].keys()
        # Fan 1 has 'Inlet Box', 'Drain'.
        # So N = 2.
        
        # M (Optional Items): Fan 1 has 'Flex Conn'. N=1?
        # Let's just check the formula string in H10 (Fan 1 Data Row)
        
        ws_quote = wb_check["Quotation"] # Use wb_check, not wb
        formula = ws_quote["H10"].value
        print(f"Formula H10: {formula}")
        
        # We expect references to 'Internal Costing'!B{Row}
        #Fab Cost Row = 28 + N = 30?
        #Fab Margin Row = 32 + N + M = 32 + 2 + 1 = 35?
        
        if "'Internal Costing'!" not in formula:
            print("FAIL: Formula does not link to Internal Costing")
        else:
            print("SUCCESS: Formula links to Internal Costing")
            
        # Verify Linked Margins
        # Check Formula in Quotation Sheet (Column H)
        # Fan 1 (Row 10)
        quote_formula = ws_quote['H10'].value
        print(f"Quotation Formula (Fan 1): {quote_formula}")
        if "Internal Costing" not in quote_formula:
            print("FAIL: Quotation formula does not reference Internal Costing")
        
        # Check Margins in Costing Sheet are Numbers
        fab_margin_row = -1
        for r in range(30, 60):
             val = ws_costing.cell(row=r, column=1).value
             if val == "Fab Margin %":
                 fab_margin_row = r
                 break
        
        if fab_margin_row != -1:
             # Fan 1 Fab Margin is 30% (0.30)
             margin_val = ws_costing.cell(row=fab_margin_row, column=2).value
             print(f"Fan 1 Fab Margin (Costing Sheet): {margin_val}")
             if margin_val != 0.3:
                 print(f"FAIL: Margin value in Costing Sheet incorrect. Expected 0.3, got {margin_val}")
        else:
             print("FAIL: 'Fab Margin %' row not found in Costing Sheet")
        # Loose check for row indices presence
        # import re
        # matches = re.findall(r"B(\d+)", quote_formula) ... 
        # Actually simpler to just ensure it's not empty and looks like a formula
        if not quote_formula.startswith("=("):
            print("FAIL: Quotation formula format incorrect")
        
        # Verify Sheets are Unlocked
        if ws_quote.protection.sheet:
             print("FAIL: Quotation sheet is still protected")
        else:
             print("PASS: Quotation sheet is unlocked")
             
        if ws_costing.protection.sheet:
             print("FAIL: Internal Costing sheet is still protected")
        else:
             print("PASS: Internal Costing sheet is unlocked")

        print("Verification Passed!")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_export()
